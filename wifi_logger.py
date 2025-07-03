import subprocess
import time
import re
from datetime import datetime

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path

LOG_FILE = "wifi_internet_issues_log.txt"
EXCEL_FILE = "wifi_internet_daily_log.xlsx"
CHECK_INTERVAL = 1  # секунд
PING_TARGET = "8.8.8.8"
PING_TIMEOUT_MS = 1000
THRESHOLD_LATENCY_MS = 500
THRESHOLD_SIGNAL_PERCENT = 30

last_state = {
    "wifi_status": None,
    "internet_status": None,
    "signal": None
}

def get_wifi_status():
    try:
        output = subprocess.check_output("netsh wlan show interfaces", shell=True, text=True)
        status = "Unknown"
        ssid = "N/A"
        signal = -1
        for line in output.splitlines():
            if "State" in line:
                status = line.split(":")[1].strip()
            elif "SSID" in line and "BSSID" not in line:
                ssid = line.split(":")[1].strip()
            elif "Signal" in line:
                signal_str = line.split(":")[1].strip().replace("%", "")
                signal = int(signal_str)
        return status, ssid, signal
    except subprocess.CalledProcessError:
        return "Error", "N/A", -1

def ping_latency(host=PING_TARGET):
    try:
        output = subprocess.check_output(f"ping -n 1 -w {PING_TIMEOUT_MS} {host}", shell=True, text=True)
        if "TTL=" in output:
            match = re.search(r"time[=<]?\s*(\d+)ms", output)
            if match:
                return int(match.group(1))
            return 0
        return -1
    except subprocess.CalledProcessError:
        return -1

def log_to_file(entry):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_entry = f"[{timestamp}] {entry}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(full_entry + "\n")

def save_to_excel(timestamp, wifi_status, ssid, signal, internet_status, latency):
    date_str = timestamp.strftime("%d.%m.%y")
    time_str = timestamp.strftime("%H:%M:%S")

    # Открываем или создаём рабочую книгу
    if Path(EXCEL_FILE).exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # Создаём лист на текущие сутки, если нет
    if date_str not in wb.sheetnames:
        ws = wb.create_sheet(title=date_str)
        headers = ["Время", "Wi-Fi статус", "SSID", "Сигнал (%)", "Интернет", "Пинг (мс)"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col)].width = 15
    else:
        ws = wb[date_str]

    # Запись строки
    ws.append([
        time_str,
        wifi_status,
        ssid,
        signal if signal >= 0 else "",
        internet_status,
        latency if latency >= 0 else ""
    ])

    # График — создаём один раз
    if not ws._charts:
        chart = LineChart()
        chart.title = "Пинг (мс) по времени"
        chart.y_axis.title = "Задержка"
        chart.x_axis.title = "Время"

        data = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 30
        chart.style = 13

        ws.add_chart(chart, "H2")

    wb.save(EXCEL_FILE)

def has_state_changed(wifi_status, internet_status, signal):
    global last_state
    changed = False
    if (wifi_status != last_state["wifi_status"] or
        internet_status != last_state["internet_status"] or
        (last_state["signal"] is not None and abs(signal - last_state["signal"]) >= 5)):
        changed = True
    last_state["wifi_status"] = wifi_status
    last_state["internet_status"] = internet_status
    last_state["signal"] = signal
    return changed

def main_loop():
    print("🔍 Старт мониторинга Wi-Fi и интернета...\n")

    while True:
        # Wi-Fi
        wifi_status, ssid, signal = get_wifi_status()
        signal_display = f"{signal}%" if signal >= 0 else "N/A"

        # Пинг
        latency = ping_latency()
        internet_status = "online" if latency >= 0 else "offline"
        ping_display = f"{latency} ms" if latency >= 0 else "timeout"

        # Формируем строку
        now_str = datetime.now().strftime("%H:%M:%S")

        status_line = (
            f"[{now_str}] Wi-Fi: {wifi_status}, "
            f"SSID: {ssid}, "
            f"Signal: {signal_display}, "
            f"Internet: {internet_status}, "
            f"Ping: {ping_display}"
        )

        print(status_line)

        # Логируем только если:
        # - состояние изменилось (Wi-Fi / Internet)
        # - плохой сигнал
        # - плохой пинг
        if (
            has_state_changed(wifi_status, internet_status, signal) or
            signal < THRESHOLD_SIGNAL_PERCENT or
            latency > THRESHOLD_LATENCY_MS
        ):
            log_to_file(status_line)
            save_to_excel(datetime.now(), wifi_status, ssid, signal, internet_status, latency)

        time.sleep(CHECK_INTERVAL)

if __name__ == "__main__":
    try:
        main_loop()
    except KeyboardInterrupt:
        print("\n🛑 Мониторинг остановлен вручную.")
