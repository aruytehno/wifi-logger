import subprocess
import unittest
from unittest.mock import patch
from datetime import datetime, timedelta
import os

import wifi_logger  # Файл с основным кодом должен быть wifi_logger.py

class TestWifiLogger(unittest.TestCase):
    """
    Набор юнит-тестов для модуля wifi_logger.

    Покрытие тестами:
    - log_to_file():
        Проверяется создание лог-файла и корректная запись текстовой строки.
    - save_to_excel():
        Проверяется:
            * создание Excel-файла;
            * создание листа с текущей датой;
            * добавление одной или нескольких строк данных;
            * корректность значений в ячейках;
            * создание линейного графика по пингу.
    - get_wifi_status():
        Через мок проверяется корректный парсинг вывода команды netsh wlan show interfaces,
        включая состояние подключения, SSID и уровень сигнала.
    - ping_latency():
        Проверяются оба случая:
            * успешный пинг с возвратом значения времени;
            * неудачный пинг с возвратом -1.
    - has_state_changed():
        Проверяется обнаружение изменения состояния Wi-Fi, интернета и уровня сигнала;
        а также отсутствие реакции на незначительные отклонения сигнала.

    Дополнительно:
    - Проверяется добавление нескольких строк в Excel, чтобы удостовериться,
      что данные не перезаписываются при повторной записи в течение суток.

    Ограничения:
    - Цикл мониторинга в main_loop() не тестируется напрямую, так как требует
      рефакторинга под тестируемые подфункции.
    """

    def setUp(self):
        """Удаляем файлы лога и Excel перед каждым тестом, чтобы тесты были изолированы."""
        self.test_time = datetime(2025, 7, 3, 12, 0, 0)
        self.test_entry = "Wi-Fi: connected, SSID: TestSSID, Signal: 99%, Internet: online, Ping: 23 ms"
        self.log_file = wifi_logger.LOG_FILE
        self.excel_file = wifi_logger.EXCEL_FILE

        for f in [self.log_file, self.excel_file]:
            if os.path.exists(f):
                os.remove(f)

    def tearDown(self):
        """Удаляем файлы лога и Excel после каждого теста для чистоты окружения."""
        for f in [self.log_file, self.excel_file]:
            if os.path.exists(f):
                os.remove(f)

    def test_log_to_file_creates_and_logs_entry(self):
        """Проверяет, что функция log_to_file создает файл и записывает в него правильную строку."""
        wifi_logger.log_to_file(self.test_entry)
        self.assertTrue(os.path.exists(self.log_file))

        with open(self.log_file, "r", encoding="utf-8") as f:
            content = f.read()
            self.assertIn(self.test_entry, content)

    def test_save_to_excel_creates_file_and_sheet(self):
        """
        Проверяет создание Excel файла, листа с правильным названием (датой),
        корректную запись данных и добавление графика.
        """
        wifi_logger.save_to_excel(
            self.test_time,
            "connected",
            "TestSSID",
            99,
            "online",
            23
        )

        self.assertTrue(os.path.exists(self.excel_file))

        from openpyxl import load_workbook
        wb = load_workbook(self.excel_file)
        sheetname = self.test_time.strftime("%d.%m.%y")
        self.assertIn(sheetname, wb.sheetnames)

        ws = wb[sheetname]

        # Проверка корректности записанных данных
        self.assertEqual(ws.cell(row=2, column=2).value, "connected")
        self.assertEqual(ws.cell(row=2, column=3).value, "TestSSID")
        self.assertEqual(ws.cell(row=2, column=4).value, 99)
        self.assertEqual(ws.cell(row=2, column=5).value, "online")
        self.assertEqual(ws.cell(row=2, column=6).value, 23)

        # Проверка, что график создан на листе
        self.assertTrue(ws._charts)

    @patch("wifi_logger.subprocess.check_output")
    def test_get_wifi_status_parsing(self, mock_subprocess):
        """
        Тестирует функцию get_wifi_status на корректный парсинг вывода команды netsh wlan show interfaces.
        Проверяет правильное извлечение статуса, SSID и уровня сигнала.
        """
        mock_subprocess.return_value = (
            "State : connected\n"
            "SSID : TestNet\n"
            "Signal : 75%"
        )
        status, ssid, signal = wifi_logger.get_wifi_status()
        self.assertEqual(status, "connected")
        self.assertEqual(ssid, "TestNet")
        self.assertEqual(signal, 75)

    @patch("wifi_logger.subprocess.check_output")
    def test_ping_latency_success(self, mock_subprocess):
        """
        Проверяет успешный разбор пинга, когда в выводе содержится TTL и время отклика.
        """
        mock_subprocess.return_value = "Reply from 8.8.8.8: bytes=32 time=42ms TTL=55"
        latency = wifi_logger.ping_latency()
        self.assertEqual(latency, 42)

    @patch("wifi_logger.subprocess.check_output", side_effect=subprocess.CalledProcessError(1, "ping"))
    def test_ping_latency_failure(self, mock_subprocess):
        """
        Проверяет, что при ошибке выполнения команды ping функция возвращает -1.
        """
        latency = wifi_logger.ping_latency()
        self.assertEqual(latency, -1)

    def test_has_state_changed_true(self):
        """
        Проверяет, что функция has_state_changed корректно выявляет изменение состояний
        Wi-Fi, интернета или значительное изменение сигнала.
        """
        wifi_logger.last_state = {
            "wifi_status": "connected",
            "internet_status": "online",
            "signal": 90
        }
        changed = wifi_logger.has_state_changed("disconnected", "offline", 50)
        self.assertTrue(changed)

    def test_has_state_changed_false(self):
        """
        Проверяет, что функция has_state_changed не отмечает изменения,
        если состояние не изменилось и изменение сигнала меньше 5%.
        """
        wifi_logger.last_state = {
            "wifi_status": "connected",
            "internet_status": "online",
            "signal": 90
        }
        changed = wifi_logger.has_state_changed("connected", "online", 92)
        self.assertFalse(changed)

    def test_multiple_entries_append(self):
        """
        Проверяет, что функция save_to_excel корректно добавляет новые строки,
        не перезаписывая существующие. После двух вызовов должны быть записаны
        две строки данных (плюс строка заголовков).
        """
        times = [self.test_time, self.test_time + timedelta(seconds=10)]
        for t in times:
            wifi_logger.save_to_excel(t, "connected", "SSID", 50, "online", 30)

        from openpyxl import load_workbook
        wb = load_workbook(self.excel_file)
        ws = wb[self.test_time.strftime("%d.%m.%y")]

        # должно быть две строки данных + заголовок
        self.assertEqual(ws.max_row, 3)


if __name__ == "__main__":
    unittest.main(verbosity=2)
